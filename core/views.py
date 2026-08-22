from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .models import Product, InventoryTransaction, Sale, SaleItem
from openpyxl import load_workbook
from django.db.models import Sum, Count, Avg, Q, F, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, time, timedelta
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.core.paginator import Paginator

def is_superuser(user):
    return user.is_authenticated and user.is_superuser

@login_required(login_url="login")
def dashboard(request):

    # =========================================================
    # DATE
    # =========================================================

    today = timezone.localdate()

    yesterday = today - timezone.timedelta(days=1)

    current_month = today.month
    current_year = today.year


    # =========================================================
    # TODAY'S SALES
    # =========================================================

    today_sales_qs = Sale.objects.filter(
        created_at__date=today
    )

    total_sales = (
        today_sales_qs.aggregate(
            total=Sum("total")
        )["total"]
        or 0
    )

    total_transactions = today_sales_qs.count()


    # =========================================================
    # ITEMS SOLD TODAY
    # =========================================================

    total_items = (
        SaleItem.objects.filter(
            sale__created_at__date=today
        ).aggregate(
            total=Sum("quantity")
        )["total"]
        or 0
    )


    # =========================================================
    # DISCOUNTS TODAY
    # =========================================================

    total_discount = (
        today_sales_qs.aggregate(
            total=Sum("discount_amount")
        )["total"]
        or 0
    )


    # =========================================================
    # INVENTORY
    # =========================================================

    total_products = Product.objects.count()

    low_stock = Product.objects.filter(
        qty__gt=0,
        qty__lte=10
    ).count()

    out_of_stock = Product.objects.filter(
        qty=0
    ).count()


    # =========================================================
    # MONTHLY SALES
    # =========================================================

    monthly_sales_qs = Sale.objects.filter(
        created_at__year=current_year,
        created_at__month=current_month
    )

    monthly_sales = (
        monthly_sales_qs.aggregate(
            total=Sum("total")
        )["total"]
        or 0
    )

    monthly_transactions = monthly_sales_qs.count()


    # =========================================================
    # YESTERDAY SALES
    # =========================================================

    yesterday_sales = (
        Sale.objects.filter(
            created_at__date=yesterday
        ).aggregate(
            total=Sum("total")
        )["total"]
        or 0
    )


    # =========================================================
    # AVERAGE SALE TODAY
    # =========================================================

    if total_transactions > 0:
        average_sale = total_sales / total_transactions
    else:
        average_sale = 0


    # =========================================================
    # TODAY VS YESTERDAY
    # =========================================================

    sales_change_percent = 0

    if yesterday_sales:
        sales_change_percent = (
            (total_sales - yesterday_sales)
            / yesterday_sales
        ) * 100


    # =========================================================
    # LAST 7 DAYS SALES DIAGRAM
    # =========================================================

    chart_labels = []
    chart_sales = []
    chart_transactions = []

    for i in range(6, -1, -1):

        chart_date = today - timezone.timedelta(days=i)

        day_qs = Sale.objects.filter(
            created_at__date=chart_date
        )

        day_total = (
            day_qs.aggregate(
                total=Sum("total")
            )["total"]
            or 0
        )

        day_transactions = day_qs.count()

        chart_labels.append(
            chart_date.strftime("%b %d")
        )

        chart_sales.append(
            float(day_total)
        )

        chart_transactions.append(
            day_transactions
        )


    # =========================================================
    # RECENT SALES
    # =========================================================

    recent_sales = (
        Sale.objects
        .select_related("staff", "discount_class")
        .prefetch_related("items")
        .order_by("-created_at")[:5]
    )


    # =========================================================
    # CONTEXT
    # =========================================================

    context = {

        # -----------------------------
        # TODAY
        # -----------------------------

        "total_sales": total_sales,

        "total_transactions": total_transactions,

        "total_items": total_items,

        "total_discount": total_discount,

        "average_sale": average_sale,


        # -----------------------------
        # INVENTORY
        # -----------------------------

        "total_products": total_products,

        "low_stock": low_stock,

        "out_of_stock": out_of_stock,


        # -----------------------------
        # PERIOD SALES
        # -----------------------------

        "monthly_sales": monthly_sales,

        "monthly_transactions": monthly_transactions,

        "yesterday_sales": yesterday_sales,

        "sales_change_percent": sales_change_percent,


        # -----------------------------
        # CHART
        # -----------------------------

        "chart_labels": chart_labels,

        "chart_sales": chart_sales,

        "chart_transactions": chart_transactions,


        # -----------------------------
        # RECENT SALES
        # -----------------------------

        "recent_sales": recent_sales,


        # -----------------------------
        # DATE
        # -----------------------------

        "today": today,
    }


    return render(
        request,
        "pages/dashboard.html",
        context
    )

@login_required
@user_passes_test(is_superuser)
def staff_list(request):

    users = User.objects.all().order_by("username")

    return render(request, "users/staff_list.html", {
        "users": users
    })

@login_required
@user_passes_test(is_superuser)
def add_staff(request):

    if request.method == "POST":

        name = request.POST.get("name", "").strip()
        username = request.POST.get("username", "").strip()
        password1 = request.POST.get("password1", "")
        password2 = request.POST.get("password2", "")

        # Checkbox
        is_staff = request.POST.get("is_staff") == "on"

        # ==============================
        # VALIDATION
        # ==============================

        if not name:
            messages.error(request, "Name is required.")
            return redirect("add_staff")

        if not username:
            messages.error(request, "Username is required.")
            return redirect("add_staff")

        if not password1:
            messages.error(request, "Password is required.")
            return redirect("add_staff")

        if password1 != password2:
            messages.error(
                request,
                "Passwords do not match."
            )
            return redirect("add_staff")

        if User.objects.filter(username=username).exists():
            messages.error(
                request,
                "Username already exists."
            )
            return redirect("add_staff")

        # ==============================
        # CREATE STAFF
        # ==============================

        user = User.objects.create(
            first_name=name,
            username=username,
            password=make_password(password1),
            is_staff=is_staff,
            is_superuser=False
        )

        messages.success(
            request,
            f"Staff account '{user.username}' created successfully."
        )

        return redirect("staff_list")

    return render(
        request,
        "users/staff_form.html"
    )


@login_required
@user_passes_test(is_superuser)
def add_staff(request):

    if request.method == "POST":

        name = request.POST.get("name", "").strip()
        username = request.POST.get("username", "").strip()
        password1 = request.POST.get("password1", "")
        password2 = request.POST.get("password2", "")

        is_staff = request.POST.get("is_staff") == "on"

        # ==============================
        # VALIDATION
        # ==============================

        if not name:
            messages.error(request, "Name is required.")
            return redirect("add-user")

        if not username:
            messages.error(request, "Username is required.")
            return redirect("add-user")

        if not password1:
            messages.error(request, "Password is required.")
            return redirect("add-user")

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect("add-user")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("add-user")

        # ==============================
        # CREATE USER
        # ==============================

        user = User.objects.create(
            first_name=name,
            username=username,
            is_staff=is_staff,
            is_superuser=False
        )

        user.set_password(password1)
        user.save()

        messages.success(
            request,
            f"Staff account '{user.username}' created successfully."
        )

        return redirect("staff_list")

    return render(
        request,
        "users/staff_form.html",
        {
            "is_edit": False,
            "user": None,
        }
    )


@login_required
@user_passes_test(is_superuser)
def edit_user(request, user_id):

    user = get_object_or_404(
        User,
        id=user_id
    )

    # ==========================================
    # PROTECT SUPERUSER
    # ==========================================

    if user.is_superuser:

        messages.error(
            request,
            "The superuser account is protected."
        )

        return redirect("staff_list")


    # ==========================================
    # POST
    # ==========================================

    if request.method == "POST":

        name = request.POST.get(
            "name",
            ""
        ).strip()

        username = request.POST.get(
            "username",
            ""
        ).strip()

        password1 = request.POST.get(
            "password1",
            ""
        )

        password2 = request.POST.get(
            "password2",
            ""
        )

        is_staff = (
            request.POST.get("is_staff") == "on"
        )


        # ==========================================
        # VALIDATE NAME
        # ==========================================

        if not name:

            messages.error(
                request,
                "Name is required."
            )

            return redirect(
                "edit-user",
                user_id=user.id
            )


        # ==========================================
        # VALIDATE USERNAME
        # ==========================================

        if not username:

            messages.error(
                request,
                "Username is required."
            )

            return redirect(
                "edit-user",
                user_id=user.id
            )


        # ==========================================
        # CHECK USERNAME
        # ==========================================

        if User.objects.filter(
            username=username
        ).exclude(
            id=user.id
        ).exists():

            messages.error(
                request,
                "Username already exists."
            )

            return redirect(
                "edit-user",
                user_id=user.id
            )


        # ==========================================
        # PASSWORD
        # ==========================================

        if password1 or password2:

            if password1 != password2:

                messages.error(
                    request,
                    "Passwords do not match."
                )

                return redirect(
                    "edit-user",
                    user_id=user.id
                )

            if len(password1) < 8:

                messages.error(
                    request,
                    "Password must be at least 8 characters."
                )

                return redirect(
                    "edit-user",
                    user_id=user.id
                )

            user.set_password(password1)


        # ==========================================
        # UPDATE USER
        # ==========================================

        user.first_name = name
        user.username = username
        user.is_staff = is_staff

        # Never allow this page to create a superuser
        user.is_superuser = False

        user.save()


        messages.success(
            request,
            f"Staff account '{user.username}' updated successfully."
        )

        return redirect("staff_list")


    # ==========================================
    # DISPLAY FORM
    # ==========================================

    return render(
        request,
        "users/staff_form.html",
        {
            "is_edit": True,
            "user": user,
        }
    )

@login_required
@user_passes_test(is_superuser)
def delete_user(request, user_id):

    user = get_object_or_404(User, id=user_id)

    # ==========================================
    # PROTECT SUPERUSER
    # ==========================================

    if user.is_superuser:

        messages.error(
            request,
            "The superuser account cannot be deleted."
        )

        return redirect("staff_list")


    # ==========================================
    # DELETE USER
    # ==========================================

    if request.method == "POST":

        username = user.username

        user.delete()

        messages.success(
            request,
            f"Staff account '{username}' deleted successfully."
        )

        return redirect("staff_list")


    # ==========================================
    # SHOW CONFIRMATION PAGE
    # ==========================================

    return render(
        request,
        "users/staff_delete.html",
        {
            "user": user
        }
    )


#Products
@login_required
@user_passes_test(is_superuser)
def product_list(request):

    products = Product.objects.all()

    return render(
        request,
        "products/product_list.html",
        {
            "products": products
        }
    )


@login_required
@user_passes_test(is_superuser)
def add_product(request):

    if request.method == "POST":

        product_model = request.POST.get(
            "product_model",
            ""
        ).strip()

        description = request.POST.get(
            "description",
            ""
        ).strip()

        category = request.POST.get(
            "category",
            ""
        ).strip()

        price = request.POST.get(
            "price",
            ""
        ).strip()

        qty = request.POST.get(
            "qty",
            ""
        ).strip()


        # ==========================================
        # VALIDATION
        # ==========================================

        if not product_model:

            messages.error(
                request,
                "Product name/model is required."
            )

            return redirect("add-product")


        if not category:

            messages.error(
                request,
                "Category is required."
            )

            return redirect("add-product")


        if not price:

            messages.error(
                request,
                "Price is required."
            )

            return redirect("add-product")


        if not qty:

            qty = 0


        # ==========================================
        # CREATE PRODUCT
        # ==========================================

        Product.objects.create(

            product_model=product_model,

            description=description,

            category=category,

            price=price,

            qty=qty

        )


        messages.success(
            request,
            f"Product '{product_model}' added successfully."
        )

        return redirect("product_list")


    return render(
        request,
        "products/product_form.html",
        {
            "is_edit": False,
            "product": None,
        }
    )


@login_required
@user_passes_test(is_superuser)
def edit_product(request, product_id):

    product = get_object_or_404(
        Product,
        id=product_id
    )


    if request.method == "POST":

        product_model = request.POST.get(
            "product_model",
            ""
        ).strip()

        description = request.POST.get(
            "description",
            ""
        ).strip()

        category = request.POST.get(
            "category",
            ""
        ).strip()

        price = request.POST.get(
            "price",
            ""
        ).strip()

        qty = request.POST.get(
            "qty",
            ""
        ).strip()


        # ==========================================
        # VALIDATION
        # ==========================================

        if not product_model:

            messages.error(
                request,
                "Product name/model is required."
            )

            return redirect(
                "edit-product",
                product_id=product.id
            )


        if not category:

            messages.error(
                request,
                "Category is required."
            )

            return redirect(
                "edit-product",
                product_id=product.id
            )


        if not price:

            messages.error(
                request,
                "Price is required."
            )

            return redirect(
                "edit-product",
                product_id=product.id
            )


        if not qty:

            qty = 0


        # ==========================================
        # UPDATE
        # ==========================================

        product.product_model = product_model

        product.description = description

        product.category = category

        product.price = price

        product.qty = qty

        product.save()


        messages.success(
            request,
            f"Product '{product.product_model}' updated successfully."
        )

        return redirect("product_list")


    return render(
        request,
        "products/product_form.html",
        {
            "is_edit": True,
            "product": product,
        }
    )


@login_required
@user_passes_test(is_superuser)
def delete_product(request, product_id):

    product = get_object_or_404(
        Product,
        id=product_id
    )


    if request.method == "POST":

        product_name = product.product_model

        product.delete()

        messages.success(
            request,
            f"Product '{product_name}' deleted successfully."
        )

        return redirect("product_list")


    return render(
        request,
        "products/product_delete.html",
        {
            "product": product
        }
    )

@login_required
@user_passes_test(is_superuser)
def import_products(request):

    if request.method == "POST":

        excel_file = request.FILES.get("excel_file")

        if not excel_file:

            messages.error(
                request,
                "Please select an Excel file."
            )

            return redirect("import-products")


        # ==========================================
        # CHECK FILE TYPE
        # ==========================================

        if not excel_file.name.lower().endswith(".xlsx"):

            messages.error(
                request,
                "Only .xlsx Excel files are allowed."
            )

            return redirect("import-products")


        try:

            workbook = load_workbook(
                excel_file,
                data_only=True
            )

            worksheet = workbook.active


            # ==========================================
            # EXPECTED HEADERS
            # ==========================================

            expected_headers = [
                "product_model",
                "description",
                "category",
                "price",
                "qty",
            ]


            headers = [
                str(cell.value).strip().lower()
                if cell.value is not None
                else ""
                for cell in worksheet[1]
            ]


            if headers != expected_headers:

                messages.error(
                    request,
                    "Invalid Excel format. Please use the provided template."
                )

                return redirect("import-products")


            # ==========================================
            # PROCESS ROWS
            # ==========================================

            created_count = 0

            error_count = 0

            errors = []


            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),
                start=2
            ):

                # Skip completely empty rows

                if not any(
                    value is not None
                    for value in row
                ):
                    continue


                product_model = (
                    str(row[0]).strip()
                    if row[0] is not None
                    else ""
                )

                description = (
                    str(row[1]).strip()
                    if row[1] is not None
                    else ""
                )

                category = (
                    str(row[2]).strip()
                    if row[2] is not None
                    else ""
                )

                price = row[3]

                qty = row[4]


                # ==========================================
                # REQUIRED FIELDS
                # ==========================================

                if not product_model:

                    errors.append(
                        f"Row {row_number}: Product Model is required."
                    )

                    error_count += 1

                    continue


                if not category:

                    errors.append(
                        f"Row {row_number}: Category is required."
                    )

                    error_count += 1

                    continue


                if price is None:

                    errors.append(
                        f"Row {row_number}: Price is required."
                    )

                    error_count += 1

                    continue


                # ==========================================
                # PRICE VALIDATION
                # ==========================================

                try:

                    price = float(price)

                    if price < 0:

                        raise ValueError

                except (ValueError, TypeError):

                    errors.append(
                        f"Row {row_number}: Invalid price."
                    )

                    error_count += 1

                    continue


                # ==========================================
                # QTY VALIDATION
                # ==========================================

                if qty is None:

                    qty = 0

                try:

                    qty = int(qty)

                    if qty < 0:

                        raise ValueError

                except (ValueError, TypeError):

                    errors.append(
                        f"Row {row_number}: Invalid quantity."
                    )

                    error_count += 1

                    continue


                # ==========================================
                # CREATE PRODUCT
                # ==========================================

                Product.objects.create(

                    product_model=product_model,

                    description=description,

                    category=category,

                    price=price,

                    qty=qty

                )

                created_count += 1


            # ==========================================
            # RESULT
            # ==========================================

            if created_count > 0:

                messages.success(
                    request,
                    f"{created_count} product(s) imported successfully."
                )


            if error_count > 0:

                for error in errors:

                    messages.error(
                        request,
                        error
                    )


            return redirect("product_list")


        except Exception as e:

            messages.error(
                request,
                f"Unable to read Excel file: {str(e)}"
            )

            return redirect("import-products")


    return render(
        request,
        "products/import_products.html"
    )


# Inventory

@login_required
@user_passes_test(is_superuser)
def inventory(request):

    products = Product.objects.all().order_by(
        "product_model"
    )


    # ==========================================
    # SEARCH
    # ==========================================

    search = request.GET.get(
        "search",
        ""
    ).strip()

    if search:

        products = products.filter(
            Q(product_model__icontains=search) |
            Q(category__icontains=search)
        )


    # ==========================================
    # CATEGORY
    # ==========================================

    selected_category = request.GET.get(
        "category",
        ""
    )


    if selected_category:

        products = products.filter(
            category=selected_category
        )


    # ==========================================
    # STATUS
    # ==========================================

    selected_status = request.GET.get(
        "stock_status",
        ""
    )


    if selected_status == "in":

        products = products.filter(
            qty__gt=10
        )


    elif selected_status == "low":

        products = products.filter(
            qty__gt=0,
            qty__lte=10
        )


    elif selected_status == "out":

        products = products.filter(
            qty=0
        )


    # ==========================================
    # STOCK VALUE
    # ==========================================

    products = products.annotate(
        stock_value=F("price") * F("qty")
    )


    # ==========================================
    # SUMMARY
    # ==========================================

    total_products = Product.objects.count()


    total_stock = Product.objects.aggregate(
        total=Coalesce(
            Sum("qty"),
            0
        )
    )["total"]


    low_stock = Product.objects.filter(
        qty__gt=0,
        qty__lte=10
    ).count()


    out_of_stock = Product.objects.filter(
        qty=0
    ).count()


    categories = Product.objects.values_list(
        "category",
        flat=True
    ).distinct().order_by(
        "category"
    )


    return render(
        request,
        "inventory/inventory.html",
        {
            "products": products,

            "total_products": total_products,

            "total_stock": total_stock,

            "low_stock": low_stock,

            "out_of_stock": out_of_stock,

            "categories": categories,

            "search": search,

            "selected_category": selected_category,

            "selected_status": selected_status,
        }
    )

@login_required
@user_passes_test(is_superuser)
def stock_in(request):

    products = Product.objects.all().order_by(
        "product_model"
    )


    if request.method == "POST":

        product_id = request.POST.get(
            "product"
        )

        quantity = request.POST.get(
            "quantity"
        )

        reference = request.POST.get(
            "reference",
            ""
        ).strip()

        notes = request.POST.get(
            "notes",
            ""
        ).strip()


        # ==========================================
        # VALIDATE PRODUCT
        # ==========================================

        try:

            product = Product.objects.get(
                id=product_id
            )

        except Product.DoesNotExist:

            messages.error(
                request,
                "Product not found."
            )

            return redirect("stock-in")


        # ==========================================
        # VALIDATE QUANTITY
        # ==========================================

        try:

            quantity = int(quantity)

        except (TypeError, ValueError):

            messages.error(
                request,
                "Please enter a valid quantity."
            )

            return redirect("stock-in")


        if quantity <= 0:

            messages.error(
                request,
                "Quantity must be greater than zero."
            )

            return redirect("stock-in")


        # ==========================================
        # UPDATE STOCK
        # ==========================================

        previous_qty = product.qty

        product.qty += quantity

        new_qty = product.qty

        product.save()


        # ==========================================
        # CREATE TRANSACTION
        # ==========================================

        InventoryTransaction.objects.create(

            product=product,

            transaction_type="IN",

            quantity=quantity,

            previous_qty=previous_qty,

            new_qty=new_qty,

            reference=reference,

            notes=notes,

            created_by=request.user

        )


        messages.success(

            request,

            f"{quantity} units added to "
            f"{product.product_model}."

        )


        return redirect("inventory")


    return render(

        request,

        "inventory/stock_in.html",

        {
            "products": products
        }

    )


@login_required
@user_passes_test(is_superuser)
def inventory_history(request):

    transactions = InventoryTransaction.objects.select_related(
        "product",
        "created_by"
    ).order_by(
        "-created_at"
    )


    # ==========================================
    # SEARCH
    # ==========================================

    search = request.GET.get(
        "search",
        ""
    ).strip()


    if search:

        transactions = transactions.filter(

            Q(product__product_model__icontains=search) |

            Q(product__category__icontains=search) |

            Q(reference__icontains=search) |

            Q(notes__icontains=search) |

            Q(created_by__username__icontains=search)

        )


    # ==========================================
    # TRANSACTION TYPE
    # ==========================================

    transaction_type = request.GET.get(
        "transaction_type",
        ""
    )


    if transaction_type:

        transactions = transactions.filter(
            transaction_type=transaction_type
        )


    # ==========================================
    # PRODUCT FILTER
    # ==========================================

    product_id = request.GET.get(
        "product",
        ""
    )


    if product_id:

        transactions = transactions.filter(
            product_id=product_id
        )


    products = Product.objects.all().order_by(
        "product_model"
    )


    # ==========================================
    # SUMMARY
    # ==========================================

    total_transactions = transactions.count()


    stock_in_count = transactions.filter(
        transaction_type="IN"
    ).count()


    stock_out_count = transactions.filter(
        transaction_type="OUT"
    ).count()


    adjustment_count = transactions.filter(
        transaction_type="ADJUSTMENT"
    ).count()


    return render(

        request,

        "inventory/inventory_history.html",

        {

            "transactions": transactions,

            "products": products,

            "total_transactions": total_transactions,

            "stock_in_count": stock_in_count,

            "stock_out_count": stock_out_count,

            "adjustment_count": adjustment_count,

            "search": search,

            "transaction_type": transaction_type,

            "selected_product": product_id,

        }

    )


@login_required
@user_passes_test(is_superuser)
def product_inventory_history(request, product_id):

    # ==========================================
    # GET TARGET PRODUCT
    # ==========================================

    product = get_object_or_404(
        Product,
        id=product_id
    )


    # ==========================================
    # GET TRANSACTIONS FOR THIS PRODUCT ONLY
    # ==========================================

    transactions = InventoryTransaction.objects.filter(
        product=product
    ).select_related(
        "created_by"
    ).order_by(
        "-created_at"
    )


    # ==========================================
    # SUMMARY
    # ==========================================

    total_transactions = transactions.count()

    stock_in_count = transactions.filter(
        transaction_type="IN"
    ).count()

    stock_out_count = transactions.filter(
        transaction_type="OUT"
    ).count()

    adjustment_count = transactions.filter(
        transaction_type="ADJUSTMENT"
    ).count()


    # ==========================================
    # TOTAL STOCK IN
    # ==========================================

    from django.db.models import Sum

    total_stock_in = (
        transactions
        .filter(transaction_type="IN")
        .aggregate(
            total=Sum("quantity")
        )["total"] or 0
    )


    # ==========================================
    # TOTAL STOCK OUT
    # ==========================================

    total_stock_out = (
        transactions
        .filter(transaction_type="OUT")
        .aggregate(
            total=Sum("quantity")
        )["total"] or 0
    )


    # ==========================================
    # RENDER
    # ==========================================

    return render(

        request,

        "inventory/product_inventory_history.html",

        {

            "product": product,

            "transactions": transactions,

            "total_transactions": total_transactions,

            "stock_in_count": stock_in_count,

            "stock_out_count": stock_out_count,

            "adjustment_count": adjustment_count,

            "total_stock_in": total_stock_in,

            "total_stock_out": total_stock_out,

        }

    )


@login_required
def admin_pos(request):

    # =========================================================
    # TODAY
    # =========================================================

    today = timezone.localdate()


    # =========================================================
    # ALL TODAY'S SALES
    #
    # We intentionally do NOT use:
    #
    #     values("salesman")
    #     salesman__id
    #     salesman__username
    #
    # because your salesman field does not allow that ORM join.
    #
    # =========================================================

    all_today_sales = list(
        Sale.objects
        .filter(
            created_at__date=today
        )
        .prefetch_related(
            "items__product"
        )
        .order_by(
            "-created_at"
        )
    )


    # =========================================================
    # HELPER: GET SALESMAN NAME
    # =========================================================

    def get_salesman_name(sale):

        try:
            salesman = sale.salesman
        except Exception:
            salesman = None


        if salesman is None:

            return "No Salesman"


        # -----------------------------------------------------
        # If salesman is a User-like object
        # -----------------------------------------------------

        if hasattr(
            salesman,
            "get_full_name"
        ):

            try:

                full_name = salesman.get_full_name().strip()

            except Exception:

                full_name = ""


            if full_name:

                return full_name


            if hasattr(
                salesman,
                "username"
            ):

                username = str(
                    salesman.username
                ).strip()

                if username:

                    return username


        # -----------------------------------------------------
        # Otherwise use the value directly
        # -----------------------------------------------------

        return str(
            salesman
        ).strip() or "No Salesman"


    # =========================================================
    # HELPER: GET PAYMENT STATUS
    # =========================================================

    def get_payment_status(sale):

        try:

            total = sale.total or 0
            payment = sale.payment or 0

        except Exception:

            total = 0
            payment = 0


        if payment >= total:

            return "PAID"

        elif payment > 0:

            return "PARTIAL"

        return "UNPAID"


    # =========================================================
    # HELPER: GET SALE BALANCE
    # =========================================================

    def get_balance(sale):

        try:

            total = sale.total or 0
            payment = sale.payment or 0

            return max(
                total - payment,
                0
            )

        except Exception:

            return 0


    # =========================================================
    # ADD DISPLAY INFORMATION TO SALES
    # =========================================================

    for sale in all_today_sales:

        sale.display_salesman = get_salesman_name(
            sale
        )

        sale.display_payment_status = get_payment_status(
            sale
        )

        sale.display_balance = get_balance(
            sale
        )


    # =========================================================
    # SUMMARY
    #
    # These are based on ALL sales today,
    # not filtered sales.
    # =========================================================

    total_sales = sum(
        (sale.total or 0)
        for sale in all_today_sales
    )


    total_transactions = len(
        all_today_sales
    )


    total_items = 0

    for sale in all_today_sales:

        try:

            for item in sale.items.all():

                total_items += (
                    item.quantity or 0
                )

        except Exception:

            pass


    total_discount = sum(
        (sale.discount_amount or 0)
        for sale in all_today_sales
    )


    total_payment = sum(
        (sale.payment or 0)
        for sale in all_today_sales
    )


    # =========================================================
    # AVERAGE TRANSACTION
    # =========================================================

    if total_transactions > 0:

        average_transaction = (
            total_sales /
            total_transactions
        )

    else:

        average_transaction = 0


    # =========================================================
    # TOTAL AMOUNT STILL DUE
    # =========================================================

    total_due = sum(
        sale.display_balance
        for sale in all_today_sales
    )


    # =========================================================
    # PAYMENT STATUS COUNTS
    # =========================================================

    paid_transactions = sum(
        1
        for sale in all_today_sales
        if sale.display_payment_status == "PAID"
    )


    partial_transactions = sum(
        1
        for sale in all_today_sales
        if sale.display_payment_status == "PARTIAL"
    )


    unpaid_transactions = sum(
        1
        for sale in all_today_sales
        if sale.display_payment_status == "UNPAID"
    )


    # =========================================================
    # SALESMAN OPTIONS
    # =========================================================

    salesman_options = sorted(
        {
            sale.display_salesman
            for sale in all_today_sales
            if sale.display_salesman
        },
        key=lambda name: name.lower()
    )


    # =========================================================
    # GET FILTERS
    # =========================================================

    search = request.GET.get(
        "search",
        ""
    ).strip()


    salesman_filter = request.GET.get(
        "salesman",
        ""
    ).strip()


    payment_filter = request.GET.get(
        "payment_status",
        ""
    ).strip().upper()


    # =========================================================
    # FILTER SALES
    #
    # Filtering is done in Python so it works regardless of
    # whether salesman is a CharField, FK, property, etc.
    # =========================================================

    filtered_sales = []


    for sale in all_today_sales:

        # -----------------------------------------------------
        # SEARCH
        # -----------------------------------------------------

        if search:

            search_lower = search.lower()


            sale_id_text = str(
                sale.id
            ).lower()


            salesman_text = (
                sale.display_salesman or ""
            ).lower()


            if (
                search_lower not in sale_id_text
                and
                search_lower not in salesman_text
            ):

                continue


        # -----------------------------------------------------
        # SALESMAN FILTER
        # -----------------------------------------------------

        if salesman_filter:

            if (
                sale.display_salesman
                != salesman_filter
            ):

                continue


        # -----------------------------------------------------
        # PAYMENT FILTER
        # -----------------------------------------------------

        if payment_filter:

            if (
                sale.display_payment_status
                != payment_filter
            ):

                continue


        filtered_sales.append(
            sale
        )


    # =========================================================
    # SALES PAGINATION
    # =========================================================

    sales_paginator = Paginator(
        filtered_sales,
        10
    )


    sales_page_number = request.GET.get(
        "page",
        1
    )


    sales_page_obj = sales_paginator.get_page(
        sales_page_number
    )


    # =========================================================
    # SALESMAN PERFORMANCE
    # =========================================================

    salesman_data = {}


    for sale in all_today_sales:

        salesman_name = (
            sale.display_salesman
        )


        if salesman_name not in salesman_data:

            salesman_data[salesman_name] = {

                "salesman":
                    salesman_name,

                "transactions":
                    0,

                "total_sales":
                    0,

                "total_paid":
                    0,

                "items_sold":
                    0,

            }


        # -----------------------------------------------------
        # Transactions
        # -----------------------------------------------------

        salesman_data[
            salesman_name
        ]["transactions"] += 1


        # -----------------------------------------------------
        # Sales
        # -----------------------------------------------------

        salesman_data[
            salesman_name
        ]["total_sales"] += (
            sale.total or 0
        )


        # -----------------------------------------------------
        # Payment
        # -----------------------------------------------------

        salesman_data[
            salesman_name
        ]["total_paid"] += (
            sale.payment or 0
        )


        # -----------------------------------------------------
        # Items
        # -----------------------------------------------------

        try:

            for item in sale.items.all():

                salesman_data[
                    salesman_name
                ]["items_sold"] += (
                    item.quantity or 0
                )

        except Exception:

            pass


    # =========================================================
    # SALESMAN LIST
    # =========================================================

    salesman_sales = list(
        salesman_data.values()
    )


    salesman_sales.sort(
        key=lambda item: item["total_sales"],
        reverse=True
    )


    # =========================================================
    # UNPAID + PARTIAL SALES
    # =========================================================

    unpaid_sales_list = [

        sale

        for sale in all_today_sales

        if sale.display_payment_status
        in [
            "UNPAID",
            "PARTIAL",
        ]

    ]


    # =========================================================
    # UNPAID TOTAL
    # =========================================================

    unpaid_total = sum(
        sale.display_balance
        for sale in unpaid_sales_list
    )


    # =========================================================
    # UNPAID PAGINATION
    # =========================================================

    unpaid_paginator = Paginator(
        unpaid_sales_list,
        10
    )


    unpaid_page_number = request.GET.get(
        "unpaid_page",
        1
    )


    unpaid_page_obj = unpaid_paginator.get_page(
        unpaid_page_number
    )


    # =========================================================
    # CONTEXT
    # =========================================================

    context = {

        # -----------------------------------------------------
        # Date
        # -----------------------------------------------------

        "today":
            today,


        # -----------------------------------------------------
        # Summary
        # -----------------------------------------------------

        "total_sales":
            total_sales,

        "total_transactions":
            total_transactions,

        "total_items":
            total_items,

        "total_discount":
            total_discount,

        "total_payment":
            total_payment,

        "average_transaction":
            average_transaction,

        "total_due":
            total_due,


        # -----------------------------------------------------
        # Payment status
        # -----------------------------------------------------

        "paid_transactions":
            paid_transactions,

        "unpaid_transactions":
            unpaid_transactions,

        "partial_transactions":
            partial_transactions,


        # -----------------------------------------------------
        # Sales
        # -----------------------------------------------------

        "recent_sales":
            sales_page_obj,

        "sales_page_obj":
            sales_page_obj,

        "sales_paginator":
            sales_paginator,


        # -----------------------------------------------------
        # Salesman
        # -----------------------------------------------------

        "salesman_sales":
            salesman_sales,

        "salesman_options":
            salesman_options,


        # -----------------------------------------------------
        # Unpaid
        # -----------------------------------------------------

        "unpaid_sales":
            unpaid_page_obj,

        "unpaid_page_obj":
            unpaid_page_obj,

        "unpaid_paginator":
            unpaid_paginator,

        "unpaid_total":
            unpaid_total,


        # -----------------------------------------------------
        # Filters
        # -----------------------------------------------------

        "search":
            search,

        "salesman_filter":
            salesman_filter,

        "payment_filter":
            payment_filter,

    }


    return render(
        request,
        "pages/pos.html",
        context
    )

@login_required
def admin_pos_detail(
    request,
    sale_id
):

    sale = get_object_or_404(
        Sale.objects
        .select_related(
            "staff",
            "discount_class"
        )
        .prefetch_related(
            "items__product"
        ),
        id=sale_id
    )

    return render(
        request,
        "admin/pos_detail.html",
        {
            "sale": sale
        }
    )


@login_required
def sales(request):

    # =========================================================
    # BASE SALES QUERY
    # =========================================================

    sales_queryset = (
        Sale.objects
        .select_related(
            "staff",
            "discount_class",
        )
        .prefetch_related(
            "items__product",
        )
        .order_by("-created_at")
    )


    # =========================================================
    # SEARCH
    # =========================================================

    search = request.GET.get("search", "").strip()

    if search:

        search_filter = (
            Q(salesman__icontains=search)
            |
            Q(created_at__icontains=search)
        )

        # Sale ID search
        if search.isdigit():

            search_filter |= Q(id=int(search))

        sales_queryset = sales_queryset.filter(search_filter)


    # =========================================================
    # SALESMAN FILTER
    #
    # IMPORTANT:
    # salesman is treated as a normal field.
    # We DO NOT use salesman__id or salesman__username.
    # =========================================================

    salesman_filter = request.GET.get("salesman", "").strip()

    if salesman_filter:

        sales_queryset = sales_queryset.filter(
            salesman=salesman_filter
        )


    # =========================================================
    # PAYMENT STATUS FILTER
    # =========================================================

    payment_status = request.GET.get(
        "payment_status",
        ""
    ).strip().upper()

    if payment_status in [
        "PAID",
        "PARTIAL",
        "UNPAID",
    ]:

        sales_queryset = sales_queryset.filter(
            payment_status=payment_status
        )


    # =========================================================
    # DATE FILTER
    # =========================================================

    date_filter = request.GET.get(
        "date",
        ""
    ).strip()

    if date_filter:

        sales_queryset = sales_queryset.filter(
            created_at__date=date_filter
        )


    # =========================================================
    # SALESMAN OPTIONS
    #
    # Since salesman is not a ForeignKey,
    # get unique values directly.
    # =========================================================

    salesman_options = (
        Sale.objects
        .exclude(
            salesman__isnull=True
        )
        .exclude(
            salesman=""
        )
        .values_list(
            "salesman",
            flat=True
        )
        .distinct()
        .order_by("salesman")
    )


    # =========================================================
    # SUMMARY
    # =========================================================

    summary = sales_queryset.aggregate(

        total_sales=Sum("total"),

        total_discount=Sum(
            "discount_amount"
        ),

        average_sale=Avg(
            "total"
        ),

        total_payment=Sum(
            "payment"
        ),
    )


    total_transactions = sales_queryset.count()


    total_sales = (
        summary["total_sales"]
        or 0
    )


    total_discount = (
        summary["total_discount"]
        or 0
    )


    average_sale = (
        summary["average_sale"]
        or 0
    )


    total_payment = (
        summary["total_payment"]
        or 0
    )


    # =========================================================
    # PAGINATION
    # =========================================================

    paginator = Paginator(
        sales_queryset,
        15
    )

    page_number = request.GET.get(
        "page"
    )

    page_obj = paginator.get_page(
        page_number
    )


    # =========================================================
    # CONTEXT
    # =========================================================

    context = {

        "sales": page_obj,

        "page_obj": page_obj,

        "paginator": paginator,

        # Summary
        "total_transactions": total_transactions,
        "total_sales": total_sales,
        "total_discount": total_discount,
        "average_sale": average_sale,
        "total_payment": total_payment,

        # Filters
        "search": search,
        "salesman_filter": salesman_filter,
        "payment_status": payment_status,
        "date_filter": date_filter,

        # Options
        "salesman_options": salesman_options,
    }


    return render(
        request,
        "pages/sales.html",
        context
    )


@login_required
def sale_detail(request, sale_id):

    sale = get_object_or_404(

        Sale.objects
        .select_related(
            "staff",
            "discount_class",
        )
        .prefetch_related(
            "items__product",
        ),

        id=sale_id
    )


    # =========================================================
    # AMOUNT DUE
    # =========================================================

    amount_due = max(
        sale.total - sale.payment,
        0
    )


    # =========================================================
    # PAYMENT STATUS
    # =========================================================

    if sale.payment >= sale.total:

        payment_status = "PAID"

    elif sale.payment > 0:

        payment_status = "PARTIAL"

    else:

        payment_status = "UNPAID"


    context = {

        "sale": sale,

        "amount_due": amount_due,

        "payment_status": payment_status,

    }


    return render(
        request,
        "pages/sale_detail.html",
        context
    )

@login_required
def reports(request):

    # =========================================================
    # CURRENT DATE / TIME
    # =========================================================
    now = timezone.localtime()

    # =========================================================
    # PERIOD
    # =========================================================
    period = request.GET.get("period", "daily")

    if period not in ["daily", "weekly", "monthly"]:
        period = "daily"

    # =========================================================
    # BASE SALES QUERY
    # =========================================================
    sales = Sale.objects.all()

    # =========================================================
    # TOTAL SALES
    # =========================================================
    total_sales = sales.aggregate(
        total=Sum("total")
    )["total"] or 0

    # =========================================================
    # TOTAL TRANSACTIONS
    # =========================================================
    total_transactions = sales.count()

    # =========================================================
    # TOTAL ITEMS SOLD
    # =========================================================
    total_items = SaleItem.objects.aggregate(
        total=Sum("quantity")
    )["total"] or 0

    # =========================================================
    # AVERAGE SALE
    # =========================================================
    average_sale = (
        total_sales / total_transactions
        if total_transactions > 0
        else 0
    )

    # =========================================================
    # CHART DATA
    # =========================================================

    if period == "daily":

        chart_queryset = (
            sales
            .annotate(period_date=TruncDay("created_at"))
            .values("period_date")
            .annotate(
                sales_total=Sum("total"),
                transaction_count=Count("id")
            )
            .order_by("period_date")
        )

    elif period == "weekly":

        chart_queryset = (
            sales
            .annotate(period_date=TruncWeek("created_at"))
            .values("period_date")
            .annotate(
                sales_total=Sum("total"),
                transaction_count=Count("id")
            )
            .order_by("period_date")
        )

    else:

        chart_queryset = (
            sales
            .annotate(period_date=TruncMonth("created_at"))
            .values("period_date")
            .annotate(
                sales_total=Sum("total"),
                transaction_count=Count("id")
            )
            .order_by("period_date")
        )

    # =========================================================
    # PREPARE CHART DATA
    # =========================================================

    chart_labels = []
    chart_sales = []
    chart_transactions = []

    for row in chart_queryset:

        date_value = row["period_date"]

        if period == "daily":
            label = timezone.localtime(date_value).strftime("%b %d")

        elif period == "weekly":
            label = timezone.localtime(date_value).strftime(
                "%b %d, %Y"
            )

        else:
            label = timezone.localtime(date_value).strftime(
                "%b %Y"
            )

        chart_labels.append(label)

        chart_sales.append(
            float(row["sales_total"] or 0)
        )

        chart_transactions.append(
            row["transaction_count"]
        )

    # =========================================================
    # RECENT SALES
    # =========================================================

    recent_sales = (
        Sale.objects
        .select_related("staff", "discount_class")
        .prefetch_related("items__product")
        .order_by("-created_at")[:10]
    )

    # =========================================================
    # CONTEXT
    # =========================================================

    context = {
        "period": period,

        "total_sales": total_sales,
        "total_transactions": total_transactions,
        "total_items": total_items,
        "average_sale": average_sale,

        "chart_labels": chart_labels,
        "chart_sales": chart_sales,
        "chart_transactions": chart_transactions,

        "recent_sales": recent_sales,

        "current_date": now,
    }

    return render(
        request,
        "pages/reports.html",
        context
    )